from openpyxl import load_workbook,Workbook
from templates import Template
class Writer:
    def use_template(self,template):
        pass
    def writeline(self,line):
        pass
class XLSXWriter(Writer):
    def __init__(self):
        self.workbook = Workbook()
        self.activesheet = self.workbook.active
        self.merges = {}
        self.cursor = [1,1]
    def use_template(self,template:Template):
        if template.filetype == "binary/xlsx":
            self.workbook = load_workbook(template.tfile)
            self.activesheet = self.workbook["Sheet1"]
            self.cursor = list(template.cursor)
            self.merges = template.merges
        else:
            raise TypeError(f"Wrong type for XLSXWriter,expected binary/xlsx,given {template.filetype}")
    def writeline(self,line,default="TRAY"):
        x = self.cursor[0]
        y = self.cursor[1]
        for value in line:
            self.activesheet.cell(y,x).value = value if value else default
            if x in self.merges:
                self.activesheet.merge_cells(start_row=y,start_column=x,end_row=y,end_column=self.merges[x])
                x = self.merges[x]
            x += 1
        self.cursor[1] += 1